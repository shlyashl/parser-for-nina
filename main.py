import json
import os
from time import sleep

from openpyxl import Workbook, load_workbook
from vk_api.exceptions import VkApiError
import vk_api


class GetVKPosts:
    def __init__(self, group_name, post_number=0, out_xlsx_name="output"):
        self.group_name = group_name
        self.api_version = "5.131"
        self.posts_at_a_time = 100
        self.post_offset = 0
        self.post_number = post_number
        self.post_filter = {
            "ad_allowed": True,
            "repost_allowed": True,
            "restricted_words": [
                "restricted_word_1"
            ]}
        self.out_xlsx_name = f"{out_xlsx_name}.xlsx"
        self.next_wb_row = 1
        self.wb, self.ws = self._get_output_wb()
        with open("config.json", "r", encoding="utf-8") as config_file:
            self.__token = json.load(config_file)["access_token"]
        self.api = vk_api.VkApi(token=self.__token, api_version="5.131")
        self.max_offset = self.get_max_offset()

    def _get_output_wb(self):
        columns = [
            "post_comments_count",
            "post_date",
            "post_type",
            "post_from_id",
            "post_likes_count",
            "post_reposts_count",
            "post_text",
            "post_owner_id",
            "post_id",
            "comment_id",
            "comment_from_id",
            "comment_text",
            "comment_date",
            "comment_level",
        ]
        if not os.path.exists(self.out_xlsx_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            wb.save(self.out_xlsx_name)
            for col, value in enumerate(columns, start=1):
                ws.cell(row=1, column=col, value=value)
            self.next_wb_row = 2
            print(f"Файл {self.out_xlsx_name} создан")
        else:
            print(f"Файл {self.out_xlsx_name} уже существует")
            wb = load_workbook(self.out_xlsx_name)
            ws = wb.active
            while ws.cell(row=self.next_wb_row, column=1).value is not None:
                self.next_wb_row += 1
        return [wb, ws]

    def write_row_to_wb(self, row):
        for col, value in enumerate(row, start=1):
            self.ws.cell(row=self.next_wb_row, column=col, value=value)
        self.wb.save(self.out_xlsx_name)
        self.next_wb_row += 1
        print(f"Данные добавлены в строку {row} в файле {self.out_xlsx_name}")

    def get_max_offset(self) -> int:
        try:
            a = self.api.method(
                method="wall.get",
                values={"domain": self.group_name, "count": self.posts_at_a_time}
            )["count"]
            return a
        except VkApiError as e:
            raise VkApiError(e)

    def parse_wall_data(self):
        data = self.api.method(method="wall.get", values={
            "domain": self.group_name,
            "offset": self.post_offset,
            "count": self.posts_at_a_time
        })["items"]

        return [
            [
                post["comments"]["count"],
                post["date"],
                post["type"],
                post["from_id"],
                post["likes"]["count"],
                post["reposts"]["count"],
                post["text"],
                post["owner_id"],
                post["id"]
            ]
            for post in data
        ]

    def parse_wall_comments(self, owner_id, post_id, comment_id=None, lvl=0):
        lvl += 1
        posts_at_a_time = 100
        comment_offset = 0
        data = []
        has_comments = 1
        while has_comments:
            data_part = self.api.method(method="wall.getComments", values={
                "owner_id": owner_id,
                "post_id": post_id,
                "count": posts_at_a_time,
                "offset": comment_offset,
                "comment_id": comment_id,
            }).get("items")
            if data_part:
                data += data_part
                comment_offset += posts_at_a_time
            else:
                has_comments = 0
        comments = []
        for comment in data:
            comments.append([
                comment["id"],
                comment["from_id"],
                comment["text"],
                comment["date"],
                lvl
            ])
            sub_comments = self.parse_wall_comments(owner_id, post_id, comment["id"], lvl)
            if sub_comments:
                comments += sub_comments
        return comments

    def run(self):
        if not self.post_number:
            self.post_number = self.get_max_offset()
        while (self.post_offset <= self.post_number) and (self.post_offset <= self.max_offset):
            posts = self.parse_wall_data()
            self.post_offset += self.posts_at_a_time
            for post_data in posts:
                if post_data[0] > 0:
                    comments = self.parse_wall_comments(post_data[-2], post_data[-1])
                    for comment in comments:
                        self.write_row_to_wb(post_data + comment)
                else:
                    self.write_row_to_wb(post_data)
                sleep(0)


if __name__ == "__main__":
    v = GetVKPosts("mossobyanin", out_xlsx_name="mossobyanin4")
    v.run()
    
